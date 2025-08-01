from aiogram import Router, F
from aiogram.types import Message
from aiogram.filters import BaseFilter
from typing import Union, List
from pathlib import Path
from create_bot import bot
from db.admin import getOwnersId
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from db.statistic import get_all_orders, get_users_stats, get_finance_stats
from aiogram.types import FSInputFile
router = Router()


class ChatTypeFilter(BaseFilter):
    def __init__(self, user_id: Union[int, List[int]]):
        self.user_ids = [user_id] if isinstance(user_id, int) else user_id or []

    async def __call__(self, message: Message) -> bool:
        return message.from_user.id in self.user_ids



@router.message(F.text == "Таблица", ChatTypeFilter(getOwnersId()))
async def handle_orders(message: Message):
    def autofit_columns(writer, sheet_name, dataframe):
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(dataframe.columns, 1):
            column = dataframe[col].astype(str)
            max_length = max(
                column.map(len).max(),
                len(str(col))
            ) + 3
            worksheet.column_dimensions[get_column_letter(idx)].width = max_length

    def colorize_boolean_columns(worksheet, dataframe, columns_to_color):
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row_idx in range(2, len(dataframe) + 2):
            for col_name in columns_to_color:
                col_idx = dataframe.columns.get_loc(col_name) + 1
                cell = worksheet.cell(row=row_idx, column=col_idx)
                value = dataframe.at[row_idx - 2, col_name]
                if value == 1:
                    cell.fill = green_fill
                    cell.value = "✅"
                else:
                    cell.fill = red_fill
                    cell.value = "❌"

    def export_to_excel(output_path):
        orders = sorted(get_all_orders(), key=lambda x: x["Id"])
        users = get_users_stats()
        finance_stats = get_finance_stats()

        orders_df = pd.DataFrame(orders)
        users_df = pd.DataFrame(users)
        monthly_df = pd.DataFrame(finance_stats["По месяцам"])

        today_dict = finance_stats["Сегодня"]
        all_time_dict = finance_stats["За всё время"]

        summary_df = pd.DataFrame([
            {
                "Период": "Сегодня",
                "Доход": today_dict["Доход (день)"],
                "Выплаты": today_dict["Выплаты (день)"],
                "Прибыль": today_dict["Прибыль (день)"]
            },
            {
                "Период": "За всё время",
                "Доход": all_time_dict["Доход (всего)"],
                "Выплаты": all_time_dict["Выплаты (всего)"],
                "Прибыль": all_time_dict["Прибыль (всего)"]
            }
        ])

        orders_df.rename(columns={
            "Id": "ID",
            "Adress": "Адрес",
            "Price": "Цена",
            "WorkerPrice": "Цена работнику",
            "WorkerId": "ID работника",
            "dateCreated": "Дата создания",
            "dateDone": "Дата завершения",
            "dateStarted": "Дата начала работы",
            "FullName": "ФИО заказчика",
            "Done": "Готов",
            "Active": "В работе"
            # "Paid" исключён полностью
        }, inplace=True)

        # Удаляем колонку "Paid", если она есть
        if "Paid" in orders_df.columns:
            orders_df.drop(columns=["Paid"], inplace=True)

        users_df.rename(columns={
            "WorkerName": "Имя работника",
            "TelegramId": "Telegram ID",
            "Status": "Статус",
            "OrdersCount": "Кол-во заказов",
            "TotalEarned": "Всего заработано"
        }, inplace=True)

        monthly_df.rename(columns={
            "Месяц": "Месяц",
            "Доход (мес)": "Доход",
            "Выплаты (мес)": "Выплаты",
            "Прибыль (мес)": "Прибыль"
        }, inplace=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            orders_df.to_excel(writer, sheet_name="Заказы", index=False)
            autofit_columns(writer, "Заказы", orders_df)
            colorize_boolean_columns(writer.sheets["Заказы"], orders_df, ["Готов", "В работе"])

            users_df.to_excel(writer, sheet_name="Работники", index=False)
            autofit_columns(writer, "Работники", users_df)

            monthly_df.to_excel(writer, sheet_name="Бухгалтерия", startrow=0, index=False)
            autofit_columns(writer, "Бухгалтерия", monthly_df)

            summary_startrow = len(monthly_df) + 3
            summary_df.to_excel(writer, sheet_name="Бухгалтерия", startrow=summary_startrow, index=False)
            autofit_columns(writer, "Бухгалтерия", summary_df)

    file_path = Path("report.xlsx")
    export_to_excel(file_path)
    document = FSInputFile(file_path)
    await bot.send_document(chat_id=message.from_user.id, document=document, caption="📊 Отчёт по заказам")
