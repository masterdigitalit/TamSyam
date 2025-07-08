import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from db.statistic import get_all_orders, get_users_stats, get_finance_stats


def autofit_columns(writer, sheet_name, dataframe):
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(dataframe.columns, 1):
        column = dataframe[col].astype(str)
        max_length = max(
            column.map(len).max(),
            len(str(col))
        ) + 3
        worksheet.column_dimensions[get_column_letter(idx)].width = max_length


def colorize_boolean_columns(worksheet, dataframe, columns_to_color):
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx in range(2, len(dataframe) + 2):
        for col_name in columns_to_color:
            col_idx = dataframe.columns.get_loc(col_name) + 1
            cell = worksheet.cell(row=row_idx, column=col_idx)
            value = dataframe.at[row_idx - 2, col_name]
            if value == 1:
                cell.fill = green_fill
                cell.value = "✅"
            else:
                cell.fill = red_fill
                cell.value = "❌"


def export_to_excel(output_path):
    orders = sorted(get_all_orders(), key=lambda x: x["Id"])
    users = get_users_stats()
    finance_stats = get_finance_stats()

    orders_df = pd.DataFrame(orders)
    users_df = pd.DataFrame(users)

    # Преобразуем месячные данные в DataFrame
    monthly_df = pd.DataFrame(finance_stats["По месяцам"])

    # Создадим DataFrame для "Сегодня" и "За всё время" — для удобства сделаем по одной строке с понятными колонками
    today_dict = finance_stats["Сегодня"]
    all_time_dict = finance_stats["За всё время"]

    # Можно объединить их в один DataFrame, например так:
    summary_df = pd.DataFrame([
        {
            "Период": "Сегодня",
            "Доход": today_dict["Доход (день)"],
            "Выплаты": today_dict["Выплаты (день)"],
            "Прибыль": today_dict["Прибыль (день)"]
        },
        {
            "Период": "За всё время",
            "Доход": all_time_dict["Доход (всего)"],
            "Выплаты": all_time_dict["Выплаты (всего)"],
            "Прибыль": all_time_dict["Прибыль (всего)"]
        }
    ])

    # Переименовываем столбцы заказов
    orders_df.rename(columns={
        "Id": "ID",
        "Adress": "Адрес",
        "Price": "Цена",
        "WorkerPrice": "Цена работнику",
        "WorkerId": "ID работника",
        "dateCreated": "Дата создания",
        "dateDone": "Дата завершения",
        "dateStarted": "Дата начала работы",
        "FullName": "ФИО заказчика",
        "Done": "Готов",
        "Active": "В работе",
        "Paid": "Оплачен работнику"
    }, inplace=True)

    # Переименовываем столбцы работников
    users_df.rename(columns={
        "WorkerName": "Имя работника",
        "TelegramId": "Telegram ID",
        "Status": "Статус",
        "OrdersCount": "Кол-во заказов",
        "TotalEarned": "Всего заработано"
    }, inplace=True)

    # Переименовываем столбцы по месяцам (чтобы было аккуратно)
    monthly_df.rename(columns={
        "Месяц": "Месяц",
        "Доход (мес)": "Доход",
        "Выплаты (мес)": "Выплаты",
        "Прибыль (мес)": "Прибыль"
    }, inplace=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Заказы
        orders_df.to_excel(writer, sheet_name="Заказы", index=False)
        autofit_columns(writer, "Заказы", orders_df)
        colorize_boolean_columns(writer.sheets["Заказы"], orders_df, ["Готов", "В работе", "Оплачен работнику"])

        # Работники
        users_df.to_excel(writer, sheet_name="Работники", index=False)
        autofit_columns(writer, "Работники", users_df)

        # Бухгалтерия
        # Сначала пишем свод по месяцам
        monthly_df.to_excel(writer, sheet_name="Бухгалтерия", startrow=0, index=False)
        autofit_columns(writer, "Бухгалтерия", monthly_df)

        # Далее ниже добавим свод по "Сегодня" и "За всё время"
        summary_startrow = len(monthly_df) + 3  # пустая строка после таблицы

        summary_df.to_excel(writer, sheet_name="Бухгалтерия", startrow=summary_startrow, index=False)
        autofit_columns(writer, "Бухгалтерия", summary_df)


# Запуск
export_to_excel("report.xlsx")
